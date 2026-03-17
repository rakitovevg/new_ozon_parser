"""Парсер листинга Ozon через уже запущенный Chrome (GUI VPS) по CDP.

Имитация человеческого поведения (скролл), селекторы из конфига.
Запускается в потоке (sync), вызывается из asyncio через run_in_executor.
"""
from __future__ import annotations

import os
import re
import time
import random
import logging
from urllib.parse import urljoin
from typing import Optional
from io import BytesIO

from app.config import (
    SEARCH_URL1,
    SEARCH_URL2,
    SELECTOR_TILE_ROOT,
    SELECTOR_PRICE,
    SELECTOR_NAME_LINK,
    SELECTOR_WAIT_TIMEOUT,
    SELECTOR_MAX_CARDS,
    REMOTE_CHROME_WS,
    USE_REMOTE_CHROME,
)

logger = logging.getLogger(__name__)

OZON_BASE_URL = "https://www.ozon.ru"


def build_search_url(brand_name: str, brand_code: str, model: str) -> str:
    """URL = URL1 + brand.name + '-' + brand.code + URL2 + model."""
    if not SEARCH_URL1 or not SEARCH_URL2:
        return ""
    name = (brand_name or "").strip().replace(" ", "-")
    code = (brand_code or "").strip()
    model_clean = (model or "").strip()
    return f"{SEARCH_URL1}{name}-{code}{SEARCH_URL2}{model_clean}"


def _build_sku_metrics_from_xlsx(xlsx_bytes: bytes) -> dict[str, dict]:
    """
    Разбирает XLSX экспорт mpstats и возвращает sku -> метрики.
    Ожидаемые столбцы (по заголовкам, на русском):
    SKU, Остаток, Выручка за 30 дней, Заказов, Рейтинг, Количество отзывов, Акция.
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    def norm(s: str | None) -> str:
        return (s or "").strip().lower()

    col_idx: dict[str, int] = {}
    for idx, name in enumerate(header_row):
        n = norm(str(name))
        if "sku" in n:
            col_idx["sku"] = idx
        elif "остаток" in n:
            col_idx["stock"] = idx
        elif "выручка" in n:
            col_idx["revenue"] = idx
        elif "заказов" in n:
            col_idx["orders"] = idx
        elif "рейтинг" in n:
            col_idx["rating"] = idx
        elif "количество отзывов" in n or "отзывов" in n:
            col_idx["reviews"] = idx
        elif "акция" in n:
            col_idx["promo"] = idx

    def to_int(val) -> int | None:
        if val is None:
            return None
        s = str(val)
        s = re.sub(r"\D", "", s)
        if not s:
            return None
        try:
            return int(s)
        except Exception:
            return None

    def to_float(val) -> float | None:
        if val is None:
            return None
        s = str(val).replace(" ", "").replace("₽", "").replace(",", ".")
        s = re.sub(r"[^\d\.]", "", s)
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None

    sku_metrics: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        sku_val = row[col_idx.get("sku", -1)] if "sku" in col_idx else None
        if not sku_val:
            continue
        sku = str(sku_val).strip()
        if not sku:
            continue
        stock = to_int(row[col_idx["stock"]]) if "stock" in col_idx else None
        revenue = to_float(row[col_idx["revenue"]]) if "revenue" in col_idx else None
        orders = to_int(row[col_idx["orders"]]) if "orders" in col_idx else None
        rating = to_float(row[col_idx["rating"]]) if "rating" in col_idx else None
        reviews = to_int(row[col_idx["reviews"]]) if "reviews" in col_idx else None
        promo = row[col_idx["promo"]] if "promo" in col_idx else None
        promo_str = str(promo).strip() if promo else None

        sku_metrics[sku] = {
            "stock": stock,
            "revenue_30d": revenue,
            "orders_30d": orders,
            "rating": rating,
            "reviews": reviews,
            "promo": promo_str or None,
        }

    logger.info("xlsx: собрали метрики для %d SKU", len(sku_metrics))
    return sku_metrics


def _parse_listing_html(html: str, min_price: float, model_filter: Optional[str], external_sku_metrics: Optional[dict[str, dict]] = None) -> list[dict]:
    """
    Parses listing HTML and returns list of {"name","price","link"}.
    Uses the same CSS selectors as Selenium-path (from config/.env).
    """
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "lxml")

    if external_sku_metrics is not None:
        sku_metrics: dict[str, dict] = external_sku_metrics
    else:
        # Сначала собираем карту SKU -> метрики из таблицы расширения (если она есть)
        # Поля: остаток, выручка, заказы, рейтинг, отзывы, акция
        sku_metrics = {}
        try:
            rows = soup.select("tr._tr_ysl04_1")
            for row in rows:
                cells = row.find_all("td")
                if len(cells) < 6:
                    continue
                # колонка с SKU — третья (индекс 2)
                sku_cell = cells[2]
                sku_span = sku_cell.find("span")
                if not sku_span:
                    continue
                sku_text = (sku_span.get_text(strip=True) or "").strip()
                if not sku_text:
                    continue
                # колонка с остатком — шестая (индекс 5)
                stock_cell = cells[5]
                stock_text = (stock_cell.get_text(" ", strip=True) or "").strip()
                # выручка за 30 дней — седьмая (индекс 6)
                revenue_cell = cells[6] if len(cells) > 6 else None
                revenue_text = (revenue_cell.get_text(" ", strip=True) if revenue_cell else "").strip()
                # заказов — восьмая (индекс 7)
                orders_cell = cells[7] if len(cells) > 7 else None
                orders_text = (orders_cell.get_text(" ", strip=True) if orders_cell else "").strip()
                # рейтинг — девятая (индекс 8)
                rating_cell = cells[8] if len(cells) > 8 else None
                rating_text = (rating_cell.get_text(" ", strip=True) if rating_cell else "").strip()
                # количество отзывов — десятая (индекс 9)
                reviews_cell = cells[9] if len(cells) > 9 else None
                reviews_text = (reviews_cell.get_text(" ", strip=True) if reviews_cell else "").strip()
                # акция — одиннадцатая (индекс 10)
                promo_cell = cells[10] if len(cells) > 10 else None
                promo_text = (promo_cell.get_text(" ", strip=True) if promo_cell else "").strip()

                def _to_int(raw: str) -> int | None:
                    raw = (raw or "").strip()
                    if not raw:
                        return None
                    try:
                        return int(re.sub(r"\D", "", raw))
                    except Exception:
                        return None

                stock_value = _to_int(stock_text)
                revenue_value = _to_int(revenue_text)
                orders_value = _to_int(orders_text)
                reviews_value = _to_int(reviews_text)

                # рейтинг может быть с точкой, поэтому отдельно
                rating_value: float | None
                try:
                    rating_clean = rating_text.replace(",", ".").strip()
                    rating_value = float(rating_clean) if rating_clean else None
                except Exception:
                    rating_value = None

                sku_metrics[sku_text] = {
                    "stock": stock_value,
                    "revenue_30d": revenue_value,
                    "orders_30d": orders_value,
                    "rating": rating_value,
                    "reviews": reviews_value,
                    "promo": promo_text or None,
                }
        except Exception:
            sku_metrics = {}

    tiles = soup.select(SELECTOR_TILE_ROOT) if SELECTOR_TILE_ROOT else []
    if not tiles:
        return []

    model_words: list[str] = []
    if model_filter:
        model_words = model_filter.lower().split(" ")

    out: list[dict] = []
    seen_links: set[str] = set()
    count = 0

    for tile in tiles[: max(1, int(SELECTOR_MAX_CARDS or 100))]:
        try:
            count += 1
            logger.info(f"tile #{count}")
            name_el = tile.select_one(SELECTOR_NAME_LINK) if SELECTOR_NAME_LINK else None
            if not name_el:
                logger.info(f"tile #{count}: пропуск — не найден SELECTOR_NAME_LINK={SELECTOR_NAME_LINK!r}")
                continue
            sibling = name_el.nextSibling
            if not sibling:
                logger.info(f"tile #{count}: пропуск — отсутствует nextSibling для ссылки")
                continue
            raw_href = sibling.get("href") if hasattr(sibling, "get") else None
            if raw_href:
                link = urljoin(OZON_BASE_URL, raw_href)
            else:
                link = raw_href or ""
                logger.info(f"link = {link}")
            # если ссылки нет, всё равно учитываем товар (link = "")
            if link:
                if link in seen_links:
                    logger.info(f"tile #{count}: пропуск — дубликат ссылки")
                    continue
                seen_links.add(link)

            # пытаемся вытащить SKU из ссылки и найти остаток в таблице расширения
            sku = None
            try:
                # берём последнюю "длинную" цифробуквенную группу из URL как SKU
                m = re.findall(r"(\d{6,})", link)
                if m:
                    sku = m[-1]
            except Exception:
                sku = None

            # даже если по SKU нет метрик, просто берём пустой словарь
            metrics = sku_metrics.get(sku) or {}

            price_el = tile.select_one(SELECTOR_PRICE) if SELECTOR_PRICE else None
            price_text = (price_el.get_text(" ", strip=True) if price_el else "") or ""
            price = int(re.sub(r"\D", "", price_text)) if price_text else 0

            # аккуратно достаём name/link из соседнего узла, он может быть None
            sibling = name_el.nextSibling
            if not sibling:
                logger.info(f"tile #{count}: пропуск — отсутствует nextSibling у name_el")
                continue
            if not hasattr(sibling, "getText"):
                logger.info(f"tile #{count}: пропуск — nextSibling без getText()")
                continue
            name = sibling.getText()

            logger.info(
                "tile #%s: price=%s, name=%s, sku=%s, model_words=%s",
                count,
                price,
                name,
                sku,
                model_words,
            )
            if price <= 0:
                logger.info(f"tile #{count}: пропуск — price <= 0 ({price})")
                continue
            if price > min_price:
                logger.info(f"tile #{count}: пропуск — price {price} > min_price {min_price}")
                continue

            if model_words:
                name_lower = name.lower()
                if not all(word in name_lower for word in model_words):
                    logger.info(f"tile #{count}: пропуск — не прошёл фильтр по model_words")
                    continue

            out.append(
                {
                    "name": name,
                    "price": price,
                    "link": link,
                    "stock": metrics.get("stock"),
                    "revenue_30d": metrics.get("revenue_30d"),
                    "orders_30d": metrics.get("orders_30d"),
                    "rating": metrics.get("rating"),
                    "reviews": metrics.get("reviews"),
                    "promo": metrics.get("promo"),
                },
            )
        except Exception as e:
            logger.exception("tile #%s: ошибка при разборе карточки: %s", count, e)
            continue

    out.sort(key=lambda x: x["price"])
    return out


def _scrape_with_remote_chrome(
    url: str,
    min_price: float,
    task_id: int,
    cancel_check_callback,
    found_products_callback,
    model_filter: Optional[str],
) -> list[dict]:
    """
    Использует уже запущенный Chrome (GUI VPS) через DevTools / CDP.
    Требует REMOTE_CHROME_WS и USE_REMOTE_CHROME=true.
    Делает переход на страницу, скроллит, забирает HTML и парсит его
    тем же кодом, что и ScrapingBee-путь (_parse_listing_html).
    """
    from playwright.sync_api import sync_playwright

    if not REMOTE_CHROME_WS:
        raise RuntimeError("REMOTE_CHROME_WS is not set")

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(REMOTE_CHROME_WS)
        # используем уже существующий контекст браузера (с профилем и расширениями),
        # чтобы не создавать "чистый" новый профиль без расширений
        if browser.contexts:
            context = browser.contexts[0]
        else:
            context = browser.new_context()
        page = context.new_page()
        # Пытаемся навигироваться 3 раза — Ozon может делать промежуточные редиректы.
        for attempt in range(3):
            try:
                page.goto(url, wait_until="networkidle", timeout=SELECTOR_WAIT_TIMEOUT * 4000)
                break
            except Exception as nav_err:
                logger.warning("remote chrome: nav attempt %s failed: %s", attempt + 1, nav_err)
                if attempt == 2:
                    browser.close()
                    logger.error("remote chrome: navigation failed after 3 attempts")
                    return []
                page.wait_for_timeout(2000)

        # большая пауза, чтобы дорисовались все динамические блоки
        page.wait_for_timeout(8000)

        # ждём появления хотя бы одной карточки (не критично, если таймаут)
        try:
            page.wait_for_selector(SELECTOR_TILE_ROOT, timeout=SELECTOR_WAIT_TIMEOUT * 1000)
        except Exception as wait_err:
            logger.warning("remote chrome: wait_for_selector timed out: %s", wait_err)

        # Пытаемся получить XLSX через кнопку "Экспорт" (к этому моменту mpstats уже инициализировался)
        sku_metrics_from_xlsx: dict[str, dict] | None = None
        try:
            btn = page.locator("text=Экспорт").first
            btn.wait_for(state="visible", timeout=SELECTOR_WAIT_TIMEOUT * 1000)
            btn.scroll_into_view_if_needed()
            with page.expect_download() as dl_info:
                btn.click()
            download = dl_info.value
            xlsx_bytes = download.content()
            sku_metrics_from_xlsx = _build_sku_metrics_from_xlsx(xlsx_bytes)
        except Exception as e:
            logger.warning("не удалось получить или разобрать XLSX экспорт: %s", e)
            sku_metrics_from_xlsx = None

        # скроллим страницу и таблицу расширения (template-id), чтобы догрузить все строки (до max_scrolls раз)
        max_cards = int(SELECTOR_MAX_CARDS or 100)
        max_scrolls = 10
        scrolls = 0
        last_html_len = 0

        while scrolls < max_scrolls:
            # общий скролл всей страницы
            page.evaluate("window.scrollBy(0, document.body.scrollHeight);")
            # отдельный “пошаговый” скролл таблицы расширения по её id="template-id"
            page.evaluate(
                """
                (function() {
                    const el = document.getElementById('template-id');
                    if (!el) return;
                    const step = el.clientHeight || 300;
                    // несколько небольших шагов, как при прокрутке колёсиком мыши
                    for (let i = 0; i < 10; i++) {
                        el.scrollTop += step;
                    }
                })();
                """
            )
            page.wait_for_timeout(random.uniform(3000, 4500))
            html_now = page.content()
            if len(html_now) == last_html_len:
                break
            last_html_len = len(html_now)
            scrolls += 1

        # забираем финальный HTML и закрываем только вкладку,
        # сам Chrome (GUI) оставляем работать
        html = page.content()
        page.close()

    # парсим HTML тем же кодом, что и раньше (ScrapingBee + Selenium путь),
    # но метрики (остаток/выручка/и т.д.) берём из XLSX, если он успешно распарсен
    found = _parse_listing_html(
        html,
        min_price=min_price,
        model_filter=model_filter,
        external_sku_metrics=sku_metrics_from_xlsx,
    )

    # нотификации и callback'и — те же, что были в ScrapingBee-пути
    for rec in found:
        if cancel_check_callback and cancel_check_callback(task_id):
            break
        name = rec["name"]
        price = rec["price"]
        stock = rec.get("stock")
        revenue = rec.get("revenue_30d")
        orders = rec.get("orders_30d")
        rating = rec.get("rating")
        reviews = rec.get("reviews")
        promo = rec.get("promo")
        link = rec["link"]
        lines = [
            "🔥 <b>Цена снижена!</b>",
            "",
            f"📦 {name}",
            f"💰 Цена: {price} ₽",
        ]
        if stock is not None:
            lines.append(f"📦 Остаток: {stock}")
        if revenue is not None:
            lines.append(f"📈 Выручка за 30д: {revenue} ₽")
        if orders is not None:
            lines.append(f"📊 Заказов за 30д: {orders}")
        if rating is not None:
            lines.append(f"⭐️ Рейтинг: {rating}")
        if reviews is not None:
            lines.append(f"💬 Отзывов: {reviews}")
        if promo:
            lines.append(f"🏷 Акция: {promo}")
        lines.append(f"🔗 <a href=\"{link}\">Купить на Ozon</a>")
        msg = "\n".join(lines)
        # send_telegram_callback пробрасывается через внешнюю функцию,
        # поэтому здесь только found_products_callback
        if found_products_callback:
            found_products_callback(rec)

    logger.info("run_parse_listing_sync (remote chrome): task_id=%s, подходящих товаров=%d", task_id, len(found))
    return found


def _get_proxy_for_selenium(proxy_url: Optional[str]) -> Optional[dict]:
    """Преобразует URL прокси в строку для --proxy-server."""
    if not proxy_url or not proxy_url.strip():
        return None
    from urllib.parse import urlparse

    raw = proxy_url.strip()
    u = urlparse(raw)

    # Если пользователь ввёл уже host:port — используем как есть
    if not u.scheme and ":" in raw and "@" not in raw:
        return {"server": raw}

    scheme = (u.scheme or "").lower()
    host = u.hostname or ""
    # для HTTP-прокси без порта по умолчанию 80, для socks5 — 1080
    port = u.port or (1080 if scheme.startswith("socks") else 80)

    if scheme.startswith("socks"):
        # Для SOCKS Chrome ожидает схему
        server = f"{scheme}://{host}:{port}"
    else:
        # Для обычного HTTP/HTTPS — только host:port
        server = f"{host}:{port}"

    return {"server": server}


def run_parse_listing_sync(
    url: str,
    min_price: float,
    proxy_url: Optional[str],
    send_telegram_callback,
    task_id: int,
    cancel_check_callback,
    found_products_callback,
    model_filter: Optional[str] = None,
) -> list[dict]:
    """
    Синхронный прогон: открывает url в уже запущенном Chrome (через CDP),
    скроллит страницу, парсит до SELECTOR_MAX_CARDS карточек.
    Для карточек с ценой <= min_price отправляет уведомление в Telegram и
    добавляет в found_products_callback.
    cancel_check_callback(task_id) -> True означает остановку.
    Возвращает список {"name", "price", "link"}.
    """
    found_products: list[dict] = []

    try:
        if USE_REMOTE_CHROME and REMOTE_CHROME_WS:
            logger.info("run_parse_listing_sync: task_id=%s using remote Chrome", task_id)
            found_products = _scrape_with_remote_chrome(
                url=url,
                min_price=min_price,
                task_id=task_id,
                cancel_check_callback=cancel_check_callback,
                found_products_callback=found_products_callback,
                model_filter=model_filter,
            )
            # Telegram уведомления — здесь, чтобы формат совпадал со старым кодом
            for rec in found_products:
                name = rec["name"]
                price = rec["price"]
                stock = rec.get("stock")
                revenue = rec.get("revenue_30d")
                orders = rec.get("orders_30d")
                rating = rec.get("rating")
                reviews = rec.get("reviews")
                promo = rec.get("promo")
                link = rec["link"]
                lines = [
                    "🔥 <b>Цена снижена!</b>",
                    "",
                    f"📦 {name}",
                    f"💰 Цена: {price} ₽",
                ]
                if stock is not None:
                    lines.append(f"📦 Остаток: {stock}")
                if revenue is not None:
                    lines.append(f"📈 Выручка за 30д: {revenue} ₽")
                if orders is not None:
                    lines.append(f"📊 Заказов за 30д: {orders}")
                if rating is not None:
                    lines.append(f"⭐️ Рейтинг: {rating}")
                if reviews is not None:
                    lines.append(f"💬 Отзывов: {reviews}")
                if promo:
                    lines.append(f"🏷 Акция: {promo}")
                lines.append(f"🔗 <a href=\"{link}\">Купить на Ozon</a>")
                msg = "\n".join(lines)
                if send_telegram_callback:
                    send_telegram_callback(msg)
            return found_products

        raise RuntimeError("REMOTE_CHROME not configured")

    except Exception as e:
        logger.exception("run_parse_listing_sync: task_id=%s error %s", task_id, e)
        raise